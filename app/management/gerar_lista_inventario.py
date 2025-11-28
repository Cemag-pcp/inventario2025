import os
from openpyxl import load_workbook
from app import db
from app.models import Local, Peca  # Ajuste o import conforme a estrutura da sua aplicação
from app.management.lista_lideres import carregar_lideres

# Criação do diretório de saída
output_dir = "listas_geradas_2025"
os.makedirs(output_dir, exist_ok=True)

# Caminho para o arquivo modelo
template_path = "excel_dados_inventario/Lista - Inventário 2025.xlsx"

def generate_inventory_lists():
    # Agrupando locais por (nome, almoxarifado)
    agrupados = {}
    locais = Local.query.order_by(Local.estante).all()

    # Agrupei dessa forma pois existem locais com o mesmo nome no mesmo almoxarifado, mas possui estantes difentes
    # então agrupei para evitar que fosse feita uma lista por estante
    for local in locais:
        chave = (local.nome, local.almoxarifado)
        if chave not in agrupados:
            agrupados[chave] = []
        agrupados[chave].extend(local.pecas)  # Adiciona as peças ao grupo correspondente

    # Gerando arquivos por grupo
    for (nome, almoxarifado), pecas in agrupados.items():
        piece_count = 0  # Contador de peças
        file_index = 1   # Índice do arquivo
        wb = None
        ws = None

        for peca in pecas:
            # Se contador é 0 ou múltiplo de 20, cria novo arquivo
            if piece_count % 20 == 0:
                if wb:
                    # Salva o arquivo anterior
                    output_file = os.path.join(output_dir, f"{nome}_{file_index}_{almoxarifado}.xlsx")
                    wb.save(output_file)
                    print(f"Arquivo gerado: {output_file}")
                    file_index += 1  # Incrementa o índice do arquivo

                # Cria um novo workbook e worksheet
                wb = load_workbook(template_path)
                ws = wb.active

                # Preenche o cabeçalho para o novo arquivo
                lider = carregar_lideres(almoxarifado, nome)
                ws["B2"] = lider  # Nome do líder
                ws["B4"] = almoxarifado  # Almoxarifado
                ws.merge_cells("B4:D4")  # Mesclar B4, C4, e D4
                ws["F4"] = nome  # Nome do Local

                start_row = 7  # Reinicia a escrita de peças na linha 7

            # Adiciona a peça na planilha
            row = start_row + (piece_count % 20)  # Calcula a linha com base no contador
            ws[f"A{row}"] = peca.codigo  # Código da peça
            ws[f"B{row}"] = peca.descricao  # Descrição da peça
            ws.merge_cells(f"B{row}:E{row}")  # Mesclar B7, C7, D7, e E7
            ws[f"G{row}"] = peca.local.estante  # Estante

            piece_count += 1  # Incrementa o contador de peças

        # Salva o último arquivo restante
        if wb:
            output_file = os.path.join(output_dir, f"{nome}_{file_index}_{almoxarifado}.xlsx")
            wb.save(output_file)
            print(f"Arquivo gerado: {output_file}")